# %% Importando el paquete openpyxl
import openpyxl

# %% Creando un libro excel
wb = openpyxl.Workbook()

# %% Acceder al nombre de la hoja activa y modificarlo
hoja = wb.active
print(f'Hoja activa: {hoja.title}')

hoja.title = "Hoja1"
print(f'Hoja activa: {wb.active.title}')

# %% Mostrando las hojas del libro
print(wb.sheetnames)

# %% Creando hojas en el libro
hoja2 = wb.create_sheet("Hoja2")
print(wb.sheetnames)

# %% Borrando hojas del libro
del wb['Hoja21']
print(wb.sheetnames)

# %% Accediendo a una hoja del libro
hoja = wb.active
print(f'Hoja activa: {hoja.title}')

hoja = wb['Hoja22']
wb.active = hoja
print(f'Hoja activa: {wb.active.title}')

# %% Acceder a una celda
hoja = wb.active
hoja["A1"] = "Gabriel"
print(hoja["A1"].value)

hoja["B1"] = "Villacis"
print(hoja["B1"].value)

c1 = hoja.cell(row=1, column=3, value=29)
print(c1.value)

# %% Guardar una lista de valores
productos = [
    ('producto_1', 'a859', 1500, 9.95),
    ('producto_2', 'b125', 600, 4.95),
    ('producto_3', 'c764', 200, 19.95),
    ('producto_4', 'd399', 2000, 49.95)
]

wb = openpyxl.Workbook()
hoja = wb.active

# Crea la fila del encabezado con los títulos
hoja.append(('Nombre', 'Referencia', 'Stock', 'Precio'))

for producto in productos:
    # producto es una tupla con los valores de un producto 
    hoja.append(producto)

# %% Guardando el libro como un archivo
wb.save('datasets/productos.xlsx')
# %%
